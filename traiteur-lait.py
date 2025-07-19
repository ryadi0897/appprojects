
from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl  # Excel engine
from kivy.app import App
from kivy.clock import mainthread
from kivy.lang import Builder
from kivy.properties import DictProperty, ListProperty, StringProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.image import Image
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import FadeTransition, Screen, ScreenManager
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput

# ------------------------------------------------------------
# Constants & configuration (mirrors the Tkinter version)
# ------------------------------------------------------------
FICHIER_PREFIXE = "controle-"
HEADER_ROW = ["LOT", "SNIT", "NORMAL", "MATIN", "MIDI", "SOIR"]
METADATA_FILE = "controles_metadata.json"

SESSIONS = ("MATIN", "MIDI", "SOIR")

KV = """
#:import dp kivy.metrics.dp

<Header@BoxLayout>:
    size_hint_y: None
    height: dp(56)
    padding: dp(8)
    spacing: dp(8)
    canvas.before:
        Color:
            rgba: .94,.65,0,1  # orange background (≈ #f0a500)
        Rectangle:
            pos: self.pos
            size: self.size
    Image:
        id: logo
        size_hint_x: None
        width: dp(140)
        allow_stretch: True
        keep_ratio: True
    Label:
        text: root.title_text
        bold: True
        color: 0,0,0,1
        font_size: '18sp'

<ControlButton@Button>:
    size_hint_y: None
    height: dp(48)
    background_normal: ''
    background_color: .94,.65,0,1  # orange
    color: 0,0,0,1
    font_size: '16sp'

<SessionButton@Button>:
    size_hint_y: None
    height: dp(48)
    background_normal: ''
    background_color: .37,.29,.25,1  # brown (≈ #6d4c41)
    color: 1,1,1,1
    font_size: '16sp'

<DataLabel@Label>:
    size_hint_x: None
    width: dp(100)
    color: 0,0,0,1
    font_size: '15sp'

<DataInput@TextInput>:
    multiline: False
    font_size: '15sp'

<ControlListScreen>:
    name: 'controls'
    BoxLayout:
        orientation: 'vertical'
        Header:
            id: header
            title_text: 'GESTION DE CONTROLE'
        ScrollView:
            GridLayout:
                id: list_container
                cols: 1
                size_hint_y: None
                height: self.minimum_height
                padding: dp(8)
                spacing: dp(8)
        BoxLayout:
            size_hint_y: None
            height: dp(64)
            padding: dp(8)
            ControlButton:
                text: 'NOUVEAU CONTROLE'
                on_release: root.new_control()

<SessionSelectScreen>:
    name: 'sessions'
    BoxLayout:
        orientation: 'vertical'
        Header:
            id: header
            title_text: root.control_title
        BoxLayout:
            orientation: 'vertical'
            padding: dp(24)
            spacing: dp(16)
            SessionButton:
                text: 'MATIN'
                on_release: root.open_session('MATIN')
            SessionButton:
                text: 'MIDI'
                on_release: root.open_session('MIDI')
            SessionButton:
                text: 'SOIR'
                on_release: root.open_session('SOIR')
            ControlButton:
                text: 'RETOUR'
                on_release: app.go_back()

<DataEntryScreen>:
    name: 'entry'
    BoxLayout:
        orientation: 'vertical'
        Header:
            id: header
            title_text: root.entry_title
        BoxLayout:
            orientation: 'vertical'
            padding: dp(24)
            spacing: dp(12)
            BoxLayout:
                DataLabel: text: 'LOT'
                DataInput: id: lot_in
            BoxLayout:
                DataLabel: text: 'SNIT'
                DataInput: id: snit_in
            BoxLayout:
                DataLabel: text: 'NORMAL'
                DataInput: id: normal_in
            BoxLayout:
                DataLabel: text: root.session_label
                DataInput: id: pl_in
            Label:
                id: status
                size_hint_y: None
                height: dp(24)
                color: .13,.55,.13,1  # success green by default
                text: ''
            ControlButton:
                text: 'AJOUTER / MAJ'
                on_release: root.submit()
            SessionButton:
                text: 'TERMINER'
                on_release: root.save_and_back()
"""

Builder.load_string(KV)


# ------------------------------------------------------------
# Helper functions (Excel + metadata)
# ------------------------------------------------------------

def excel_path(base_dir: Path, filename: str) -> Path:
    return base_dir / filename


def lister_fichiers_controles(base_dir: Path) -> list[str]:
    return sorted(
        [f.name for f in base_dir.glob(f"{FICHIER_PREFIXE}*.xlsx")],
        reverse=True,
    )


def charger_ou_creer_fichier_excel(path: Path) -> list[dict]:
    if not path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADER_ROW)
        wb.save(path)
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    data: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append({h: v for h, v in zip(HEADER_ROW, row)})
    return data


def sauvegarder_donnees_excel(path: Path, data: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER_ROW)
    for row in data:
        ws.append([row.get(h) for h in HEADER_ROW])
    wb.save(path)


# ------------------------------------------------------------
# Screens
# ------------------------------------------------------------
class ControlListScreen(Screen):
    def on_pre_enter(self):
        self.refresh()

    def refresh(self):
        container = self.ids.list_container
        container.clear_widgets()
        files = lister_fichiers_controles(Path(App.get_running_app().user_data_dir))
        if not files:
            container.add_widget(Label(text="Aucun contrôle trouvé", size_hint_y=None, height=40))
        for fname in files:
            btn = Button(
                text=fname.replace(FICHIER_PREFIXE, "").replace(".xlsx", ""),
                size_hint_y=None,
                height=48,
                background_normal='',
                background_color=(0.94, 0.65, 0, 1),
                color=(0, 0, 0, 1),
                on_release=lambda b, f=fname: self.open_control(f),
            )
            container.add_widget(btn)

    def open_control(self, filename: str):
        app = App.get_running_app()
        app.current_excel = filename
        app.sm.current = 'sessions'

    def new_control(self):
        now_fname = datetime.now().strftime(f"{FICHIER_PREFIXE}%m-%y.xlsx")
        app = App.get_running_app()
        app.current_excel = now_fname
        app.sm.current = 'sessions'


class SessionSelectScreen(Screen):
    control_title = StringProperty("")

    def on_pre_enter(self):
        app = App.get_running_app()
        self.control_title = app.current_excel.replace(FICHIER_PREFIXE, "").replace(".xlsx", "")

    def open_session(self, session: str):
        app = App.get_running_app()
        app.current_session = session
        app.sm.current = 'entry'


class DataEntryScreen(Screen):
    entry_title = StringProperty("")
    session_label = StringProperty("")

    # hold current workbook data in memory
    _data = ListProperty()

    def on_pre_enter(self):
        app = App.get_running_app()
        self.entry_title = f"SESSION {app.current_session}"
        self.session_label = f"PL ({app.current_session})"
        path = excel_path(Path(app.user_data_dir), app.current_excel)
        self._data = charger_ou_creer_fichier_excel(path)
        # reset form
        for wid in (self.ids.lot_in, self.ids.snit_in, self.ids.normal_in, self.ids.pl_in):
            wid.text = ""
        self.ids.status.text = ""
        self.ids.lot_in.focus = True

    def _find_row(self, lot: str, snit: int, normal: int):
        # exact or swapped snit/normal
        for row in self._data:
            cond1 = row.get('LOT') == lot
            cond2 = (row.get('SNIT') == snit and row.get('NORMAL') == normal) or (
                row.get('SNIT') == normal and row.get('NORMAL') == snit)
            if cond1 and cond2:
                return row
        return None

    def submit(self):
        lot = self.ids.lot_in.text.strip()
        snit_s = self.ids.snit_in.text.strip()
        normal_s = self.ids.normal_in.text.strip()
        pl_s = self.ids.pl_in.text.strip().replace(',', '.')

        if not all((lot, snit_s, normal_s, pl_s)):
            self._set_status("Tous les champs sont requis", error=True)
            return
        try:
            snit = int(snit_s)
            normal = int(normal_s)
            pl = float(pl_s)
        except ValueError:
            self._set_status("SNIT, NORMAL et PL doivent être numériques", error=True)
            return

        row = self._find_row(lot, snit, normal)
        if row:
            row[App.get_running_app().current_session] = pl
            self._set_status(f"Mise à jour vache {row['SNIT']}/{row['NORMAL']}")
        else:
            new_row = {h: None for h in HEADER_ROW}
            new_row.update({
                'LOT': lot,
                'SNIT': snit,
                'NORMAL': normal,
                App.get_running_app().current_session: pl,
            })
            self._data.append(new_row)
            self._set_status(f"Ajouté vache {snit}/{normal}")

        # clear inputs except LOT for faster entry
        for w in (self.ids.snit_in, self.ids.normal_in, self.ids.pl_in):
            w.text = ""
        self.ids.snit_in.focus = True

    def _set_status(self, msg: str, *, error: bool = False):
        self.ids.status.color = (0.75, 0, 0, 1) if error else (0.13, 0.55, 0.13, 1)
        self.ids.status.text = msg

    def save_and_back(self):
        app = App.get_running_app()
        path = excel_path(Path(app.user_data_dir), app.current_excel)
        try:
            sauvegarder_donnees_excel(path, self._data)
            self._popup("Succès", f"Données sauvegardées dans {app.current_excel}")
        except PermissionError:
            self._popup("Erreur", "Impossible de sauvegarder. Fichier ouvert ailleurs ?")
        app.go_back(to='sessions')

    @mainthread
    def _popup(self, title: str, text: str):
        popup = Popup(title=title, content=Label(text=text), size_hint=(0.8, 0.3))
        popup.open()


# ------------------------------------------------------------
# App class
# ------------------------------------------------------------
class SeiglaApp(App):
    current_excel: str = ""
    current_session: str = ""

    def build(self):
        icon_path = os.path.join(self.directory, 'icon.png')
        if os.path.exists(icon_path):
            self.icon = icon_path
        self.title = "SEIGLA – Gestion de Contrôle"

        self.sm = ScreenManager(transition=FadeTransition())
        self.sm.add_widget(ControlListScreen())
        self.sm.add_widget(SessionSelectScreen())
        self.sm.add_widget(DataEntryScreen())
        return self.sm

    def go_back(self, *, to: str | None = None):
        if to:
            self.sm.current = to
        else:
            self.sm.current = 'controls'


if __name__ == '__main__':
    SeiglaApp().run()
